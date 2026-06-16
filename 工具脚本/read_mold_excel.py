"""
模具报价 Excel 批量读取脚本
用法: python3 read_mold_excel.py <报价文件夹路径>
输出: 同目录下生成 分析报告.md
"""

import sys
import openpyxl
from pathlib import Path


def extract_folder(folder_path: str, output_name: str = '分析报告.md'):
    base = Path(folder_path)
    output = base / output_name

    excel_files = sorted(base.glob('*.xlsx'))
    if not excel_files:
        print(f'未找到 Excel 文件: {base}')
        return

    lines = []
    lines.append(f'# 模具报价数据提取报告 — {base.name}\n')
    lines.append(f'**文件数量**: {len(excel_files)} 个 Excel 文件\n')

    for fpath in excel_files:
        lines.append(f'\n---\n\n## 文件: {fpath.name}\n')
        try:
            wb = openpyxl.load_workbook(str(fpath), data_only=True)
        except Exception as e:
            lines.append(f'> 读取失败: {e}\n')
            continue

        lines.append(f'**Sheets**: {", ".join(wb.sheetnames)}\n')

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            lines.append(f'\n### Sheet: {sheet_name}\n')
            lines.append('| 列位 | 内容 |')
            lines.append('|------|------|')
            for row in ws.iter_rows(values_only=True):
                cells = [(i, v) for i, v in enumerate(row) if v is not None]
                if not cells:
                    continue
                row_parts = []
                for i, v in cells:
                    val = str(v).replace('\n', ' / ').replace('|', '\\|').strip()
                    if val:
                        row_parts.append(f'[{i}] {val}')
                if row_parts:
                    lines.append(f'| | {" | ".join(row_parts)} |')

    output.write_text('\n'.join(lines), encoding='utf-8')
    print(f'完成，输出文件: {output}')
    print(f'共 {len(lines)} 行')


if __name__ == '__main__':
    folder = sys.argv[1] if len(sys.argv) > 1 else '.'
    extract_folder(folder)
