#!/usr/bin/env python3
"""
特殊特性批量覆盖 - 将参考文件的检验频次覆盖到目标文件。

Usage:
  python3 batch_update_special_chars.py <参考文件.xlsx> <目标文件1.xlsx> [<目标文件2.xlsx> ...]
  python3 batch_update_special_chars.py <参考文件.xlsx> --dir <目录>
  python3 batch_update_special_chars.py <参考文件.xlsx>   # 自动扫描同目录

逻辑:
  1. 读参考文件，按 (sheet名, 特性名) 提取当前检验频次值
  2. 对每个目标文件，找同名 sheet 中同名特性行，将频次改为参考值
  3. 打印 diff 报告，只改不同的值
"""

import sys
import os
from pathlib import Path
import openpyxl

# 主要工作 sheet 及其 (特性名列, 检验频次列) - Excel 1-indexed 列号
SHEET_CONFIG = {
    '成品 特性':  (3, 8),
    '原材料特性': (3, 8),
    '过程特性':   (3, 8),
}

# 跳过这些特性名（表头行或更新履历行）
SKIP_FEATURE_NAMES = {
    '特性特性', '产品特性名称', '特性描述', '序号', '序号 ',
    'Chaiacterization/特性描述', 'Name/名称', '名称',
    '更新内容',
}

# 跳过这些频次值（表头行）
SKIP_FREQ_VALUES = {
    '测试数量', '检验或试验频次', '检验频次',
    'Inspection or Testing Frequency',
    '版本号',
}


def load_freq_map(ref_path):
    """从参考文件提取 {sheet: {feature_name: new_freq}} 映射"""
    wb = openpyxl.load_workbook(ref_path, data_only=True)
    freq_map = {}

    for sheet_name, (feat_col, freq_col) in SHEET_CONFIG.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        sheet_map = {}
        for r in range(1, ws.max_row + 1):
            feat = ws.cell(r, feat_col).value
            freq = ws.cell(r, freq_col).value
            if not feat or not isinstance(feat, str):
                continue
            feat = feat.strip()
            if feat in SKIP_FEATURE_NAMES or not feat:
                continue
            if not freq or not isinstance(freq, str):
                continue
            if freq.strip() in SKIP_FREQ_VALUES:
                continue
            sheet_map[feat] = freq.strip()
        freq_map[sheet_name] = sheet_map

    return freq_map


def apply_to_file(ref_freq_map, target_path):
    """将参考频次应用到目标文件，返回变更列表 [(sheet, row, feat, old, new)]"""
    wb = openpyxl.load_workbook(target_path)
    changes = []

    for sheet_name, (feat_col, freq_col) in SHEET_CONFIG.items():
        if sheet_name not in wb.sheetnames or sheet_name not in ref_freq_map:
            continue
        ws = wb[sheet_name]
        ref_map = ref_freq_map[sheet_name]

        for r in range(1, ws.max_row + 1):
            feat = ws.cell(r, feat_col).value
            if not feat or not isinstance(feat, str):
                continue
            feat_name = feat.strip()
            if feat_name not in ref_map:
                continue
            old_freq = ws.cell(r, freq_col).value
            new_freq = ref_map[feat_name]
            old_str = old_freq.strip() if isinstance(old_freq, str) else str(old_freq)
            if old_str != new_freq:
                ws.cell(r, freq_col).value = new_freq
                changes.append((sheet_name, r, feat_name, old_str, new_freq))

    if changes:
        wb.save(target_path)

    return changes


def main():
    args = sys.argv[1:]
    if not args:
        print("用法: python3 batch_update_special_chars.py <参考文件> [目标文件...或 --dir <目录>]")
        sys.exit(1)

    ref_path = args[0]
    ref_resolved = str(Path(ref_path).resolve())

    # 确定目标文件列表
    target_paths = []
    def valid_xlsx(p):
        return not p.name.startswith('~$')

    if len(args) == 1:
        # 自动扫描同目录
        ref_dir = Path(ref_path).parent
        target_paths = [
            str(p) for p in ref_dir.glob('*.xlsx')
            if str(p.resolve()) != ref_resolved and valid_xlsx(p)
        ]
    elif '--dir' in args:
        dir_idx = args.index('--dir')
        target_dir = Path(args[dir_idx + 1])
        target_paths = [
            str(p) for p in target_dir.glob('*.xlsx')
            if str(p.resolve()) != ref_resolved and valid_xlsx(p)
        ]
    else:
        target_paths = [p for p in args[1:] if valid_xlsx(Path(p))]

    if not target_paths:
        print("没有找到目标文件。")
        sys.exit(1)

    print(f"参考文件: {ref_path}")
    print()

    # 加载参考频次映射
    ref_freq_map = load_freq_map(ref_path)
    print("参考文件检验频次映射:")
    for sheet, mapping in ref_freq_map.items():
        for feat, freq in mapping.items():
            print(f"  [{sheet}] {feat} → {freq}")
    print()

    # 依次覆盖每个目标文件
    total_changes = 0
    for target_path in sorted(target_paths):
        target_name = os.path.basename(target_path)
        try:
            changes = apply_to_file(ref_freq_map, target_path)
        except Exception as e:
            print(f"❌ {target_name}  错误: {e}")
            continue

        if changes:
            print(f"✅ {target_name}  ({len(changes)} 处修改)")
            for sheet, row, feat, old, new in changes:
                print(f"   [{sheet}] 行{row} 「{feat}」: {old!r} → {new!r}")
            total_changes += len(changes)
        else:
            print(f"⚪ {target_name}  (无需修改或无匹配特性)")

    print(f"\n合计修改 {total_changes} 处，共 {len(target_paths)} 个目标文件。")


if __name__ == '__main__':
    main()
