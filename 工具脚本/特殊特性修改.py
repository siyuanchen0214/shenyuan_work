#!/usr/bin/env python3
"""
特殊特性修改.py — 三步一键：同步 body → 导出 PDF → 合并打印文件

用法:
    python3 特殊特性修改.py <文件夹路径> <参考文件名.xlsx>

步骤:
    1. 只读参考文件，提取三个主 sheet 的数据区（检验频次、备注等所有列）
    2. 将目标文件夹内其他 xlsx 的数据区对齐到参考文件（不碰表头/零件号区域）
    3. 将参考文件 + 所有目标文件依次导出横向单页 PDF，合并为一个文件
       输出: <目标文件夹>/特殊特性合并_YYYYMMDD.pdf
"""

import sys
import os
import glob
import tempfile
from copy import copy
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from openpyxl.utils import get_column_letter

try:
    import xlwings as xw
except ImportError:
    sys.exit("缺少 xlwings，请先运行: pip3 install xlwings")

try:
    from pypdf import PdfReader, PdfWriter
except ImportError:
    sys.exit("缺少 pypdf，请先运行: pip3 install pypdf")


# ─── 配置 ───────────────────────────────────────────────────────────────────
MAIN_SHEETS = ['成品 特性', '原材料特性', '过程特性']
# 数据区要同步的列范围（1-indexed Excel 列号，C=3 到 K=11）
SYNC_COL_START = 3   # C
SYNC_COL_END   = 11  # K


# ─── Step 1: 读参考文件 body（只读）─────────────────────────────────────────

def find_data_range(ws):
    """找到数据行范围：表头行之后 → 更新履历之前"""
    header_row = None
    end_row = ws.max_row
    for r in range(1, ws.max_row + 1):
        b = ws.cell(r, 2).value  # col B
        if b and isinstance(b, str) and '序号' in b:
            header_row = r
        c = ws.cell(r, 2).value
        if c and isinstance(c, str) and '更新履历' in c:
            end_row = r - 1
            break
    data_start = (header_row + 1) if header_row else 8
    return data_start, end_row


def read_reference_body(ref_path):
    """
    从参考文件提取各主 sheet 的数据区（值 + 格式）。
    返回: {sheet_name: {'start': int, 'end': int, 'rows': {row: {col: {value, alignment, font, fill, border, number_format}}}}}
    只读，不保存文件。
    """
    wb = openpyxl.load_workbook(ref_path)  # 不用 data_only，才能读到格式
    body = {}
    for sh in MAIN_SHEETS:
        if sh not in wb.sheetnames:
            continue
        ws = wb[sh]
        data_start, data_end = find_data_range(ws)
        rows = {}
        for r in range(data_start, data_end + 1):
            row_data = {}
            for c in range(SYNC_COL_START, SYNC_COL_END + 1):
                cell = ws.cell(r, c)
                row_data[c] = {
                    'value':         cell.value,
                    'alignment':     copy(cell.alignment),
                    'font':          copy(cell.font),
                    'fill':          copy(cell.fill),
                    'border':        copy(cell.border),
                    'number_format': cell.number_format,
                }
            rows[r] = row_data
        body[sh] = {'start': data_start, 'end': data_end, 'rows': rows}
        print(f"    [{sh}] 数据行 {data_start}–{data_end}，共 {data_end - data_start + 1} 行")
    return body


# ─── Step 2: 同步 body 到目标文件 ────────────────────────────────────────────

def apply_body_to_file(body, target_path):
    """将参考 body（值 + 格式）写入目标文件，返回修改次数。"""
    wb = openpyxl.load_workbook(target_path)
    total_changes = 0
    for sh, info in body.items():
        if sh not in wb.sheetnames:
            continue
        ws = wb[sh]
        for r, col_vals in info['rows'].items():
            for c, ref in col_vals.items():
                cell = ws.cell(r, c)
                ref_val = ref['value']

                # 值比较
                old_s = cell.value.strip() if isinstance(cell.value, str) else cell.value
                new_s = ref_val.strip()    if isinstance(ref_val,    str) else ref_val
                val_changed = (old_s != new_s)

                # 格式比较（wrap_text 为主，其余也一并对齐）
                fmt_changed = (
                    cell.alignment.wrap_text != ref['alignment'].wrap_text or
                    cell.font.bold           != ref['font'].bold           or
                    cell.number_format       != ref['number_format']
                )

                if val_changed or fmt_changed:
                    if val_changed:
                        cell.value = ref_val
                    cell.alignment     = copy(ref['alignment'])
                    cell.font          = copy(ref['font'])
                    cell.fill          = copy(ref['fill'])
                    cell.border        = copy(ref['border'])
                    cell.number_format = ref['number_format']
                    total_changes += 1

    wb.save(target_path)
    return total_changes


# ─── Step 3: 导出 PDF + 合并 ─────────────────────────────────────────────────

def get_content_range(ws):
    min_row = min_col = float('inf')
    max_row = max_col = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                min_row = min(min_row, cell.row)
                min_col = min(min_col, cell.column)
                max_row = max(max_row, cell.row)
                max_col = max(max_col, cell.column)
    for merge in ws.merged_cells.ranges:
        min_row = min(min_row, merge.min_row)
        min_col = min(min_col, merge.min_col)
        max_row = max(max_row, merge.max_row)
        max_col = max(max_col, merge.max_col)
    if max_row == 0:
        return None
    return min_row, min_col, max_row, max_col


def prepare_for_pdf(xlsx_path, tmp_path):
    """复制到临时文件并设置横向 fit-to-page，返回可见 tab 数。"""
    wb = openpyxl.load_workbook(xlsx_path)
    visible = 0
    for ws in wb.worksheets:
        if ws.sheet_state == 'hidden':
            continue
        visible += 1
        rng = get_content_range(ws)
        if rng is None:
            continue
        min_row, min_col, max_row, max_col = rng
        min_c = get_column_letter(min_col)
        max_c = get_column_letter(max_col)
        ws.print_area = f"${min_c}${min_row}:${max_c}${max_row}"
        ws.page_setup.orientation = 'landscape'
        ws.page_setup.fitToWidth  = 1
        ws.page_setup.fitToHeight = 1
        if ws.sheet_properties is None:
            ws.sheet_properties = WorksheetProperties()
        if ws.sheet_properties.pageSetUpPr is None:
            ws.sheet_properties.pageSetUpPr = PageSetupProperties()
        ws.sheet_properties.pageSetUpPr.fitToPage = True
    wb.save(tmp_path)
    return visible


def convert_to_pdf(xlsx_path, pdf_path, app):
    """用 xlwings 把单个 xlsx 转为 PDF，返回页数。"""
    tmp_fd, tmp_path = tempfile.mkstemp(suffix='.xlsx', prefix='_tmp_')
    os.close(tmp_fd)
    try:
        visible_count = prepare_for_pdf(xlsx_path, tmp_path)
        wb = app.books.open(str(Path(tmp_path).resolve()))
        wb.to_pdf(path=str(Path(pdf_path).resolve()))
        wb.close()
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)
    pdf_pages = len(PdfReader(pdf_path).pages)
    return pdf_pages, visible_count


def merge_pdfs(pdf_paths, output_path):
    """合并多个 PDF 文件为一个。"""
    writer = PdfWriter()
    for p in pdf_paths:
        for page in PdfReader(p).pages:
            writer.add_page(page)
    with open(output_path, 'wb') as f:
        writer.write(f)
    return len(PdfReader(output_path).pages)


# ─── 主流程 ──────────────────────────────────────────────────────────────────

def main():
    if len(sys.argv) < 3:
        print(__doc__)
        sys.exit(1)

    target_dir = sys.argv[1]
    ref_name   = sys.argv[2]

    if not os.path.isdir(target_dir):
        sys.exit(f"文件夹不存在: {target_dir}")

    ref_path = os.path.join(target_dir, ref_name)
    if not os.path.isfile(ref_path):
        sys.exit(f"在文件夹内找不到参考文件: {ref_name}")

    ref_resolved = str(Path(ref_path).resolve())

    # 收集目标文件夹内所有 xlsx（排除参考文件和临时锁文件）
    all_xlsx = sorted([
        p for p in glob.glob(os.path.join(target_dir, '*.xlsx'))
        if not os.path.basename(p).startswith('~$')
    ])
    target_xlsx = [p for p in all_xlsx if str(Path(p).resolve()) != ref_resolved]

    print(f"\n参考文件: {os.path.basename(ref_path)}")
    print(f"目标文件: {len(target_xlsx)} 个\n")

    # ── 步骤 1：读参考 body ──────────────────────────────────────────────────
    print("【步骤 1】读取参考文件 body（只读）...")
    body = read_reference_body(ref_path)
    print()

    # ── 步骤 2：同步目标文件 ──────────────────────────────────────────────────
    print("【步骤 2】同步目标文件...")
    for p in target_xlsx:
        name = os.path.basename(p)
        changes = apply_body_to_file(body, p)
        status = f"{changes} 处更新" if changes > 0 else "无需修改"
        print(f"  ✅ {name}  ({status})")
    print()

    # ── 步骤 3：导出 PDF + 合并 ──────────────────────────────────────────────
    print("【步骤 3】导出 PDF...")
    tmp_dir = tempfile.mkdtemp(prefix='特殊特性_pdf_')
    pdf_paths = []

    # 参考文件排第一，其余按文件名排序
    ordered_xlsx = [ref_path] + target_xlsx

    app = xw.App(visible=False)
    try:
        for xlsx in ordered_xlsx:
            name = os.path.basename(xlsx)
            pdf_out = os.path.join(tmp_dir, os.path.splitext(name)[0] + '.pdf')
            print(f"  转换: {name}")
            try:
                pages, tabs = convert_to_pdf(xlsx, pdf_out, app)
                status = f"✓ {pages}页/{tabs}tab"
            except Exception as e:
                print(f"    ❌ 出错: {e}")
                continue
            print(f"         {status}")
            pdf_paths.append(pdf_out)
    finally:
        app.quit()

    # 合并
    timestamp   = datetime.now().strftime('%Y%m%d')
    merged_name = f"特殊特性合并_{timestamp}.pdf"
    merged_path = os.path.join(target_dir, merged_name)
    total_pages = merge_pdfs(pdf_paths, merged_path)

    print(f"\n合并完成: {merged_name}")
    print(f"总页数:   {total_pages} 页（{len(pdf_paths)} 个文件）")
    print(f"路径:     {merged_path}\n")


if __name__ == '__main__':
    main()
