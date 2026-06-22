#!/usr/bin/env python3
# 生成「塑料原料牌号汇总」Excel —— 两个tab,可长期累积。
#   Tab1 注塑原材料: 我们按kg采购、自制注塑成型的塑料原料(逐零件一行)
#   Tab2 总成其余子件: 采购件/压制件/纺织件/金属标准件等
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = "/Users/chensiyuan5/Desktop/Github/shenyuan_work/注塑项目/塑料原料牌号汇总.xlsx"

# ---------- Tab1: 注塑原材料 ----------
COLS1 = ["项目", "所属总成", "零件名称(中)", "零件号", "材料大类",
         "原材料牌号/材料规格", "厂家/供应商", "颜色", "数据来源文件", "备注"]
rows1 = [
    # 汽车大灯 FW_P06 (客户:迅驰/朗恩斯)
    ["汽车大灯 FW_P06", "—", "远光灯灯光导", "—", "PC 聚碳酸酯(光学级)",
     "PC LED2245 000000", "科思创 Covestro", "无色透明",
     "远光灯灯光导.xlsx (模具式样书 4-SJ-074)", "光学件,双面3mic镜面抛光"],
    ["汽车大灯 FW_P06", "—", "近光灯灯光导", "—", "PC 聚碳酸酯(光学级)",
     "PC LED2245 000000", "科思创 Covestro", "无色透明",
     "近光灯灯光导.xlsx (模具式样书 4-SJ-074)", "光学件"],
    ["汽车大灯 FW_P06", "—", "前雾灯光导", "—", "PC 聚碳酸酯(光学级)",
     "PC LED2245 000000", "科思创 Covestro", "无色透明",
     "P06前雾灯 模具式样书...V1.0(解密).xlsx (sheet 前灯双色灯罩)", "光学件"],
    ["汽车大灯 FW_P06", "—", "远光辅助光导", "—", "PC 聚碳酸酯(光学级)",
     "PC LED2245 000000", "科思创 Covestro", "无色透明",
     "远光辅助光导.xlsx (模具式样书 4-SJ-074)", "光学件;STP仅几何无材料"],
    # GEA 轮罩 (客户:上海金榕/Renault RG4F1) —— 气坝SPOILER注塑件
    ["GEA 轮罩", "左前挡泥皮总成 638431477R (PHEV/HEV)", "气坝前左 SPOILER-FR LH", "638431477R", "TPV 热塑性硫化橡胶",
     "Santoprene 103-40 (PMR:UB02b,UB05d)", "CELANESE 塞拉尼斯", "TBD",
     "技术协议_20260615.docx; Part List; 报价表-深远", "密度0.9"],
    ["GEA 轮罩", "右前挡泥皮总成 638422166R (PHEV/HEV)", "气坝前右 SPOILER-FR RH", "638422166R", "TPV",
     "Santoprene 103-40 (PMR:UB02b,UB05d)", "CELANESE 塞拉尼斯", "TBD",
     "技术协议_20260615.docx; Part List; 报价表-深远", "密度0.9"],
    ["GEA 轮罩", "左前挡泥皮总成 638439763R (BEV)", "气坝前左 SPOILER-FR LH", "638439763R", "TPV",
     "Santoprene 103-40 (PMR:UB02b,UB05d)", "CELANESE 塞拉尼斯", "TBD",
     "技术协议_20260615.docx; Part List; 报价表-深远", "密度0.9"],
    ["GEA 轮罩", "右前挡泥皮总成 638424366R (BEV)", "气坝前右 SPOILER-FR RH", "638424366R", "TPV",
     "Santoprene 103-40 (PMR:UB02b,UB05d)", "CELANESE 塞拉尼斯", "TBD",
     "技术协议_20260615.docx; Part List; 报价表-深远", "密度0.9"],
    ["GEA 轮罩", "后左挡泥皮总成 767496296R", "气坝后左 SPOILER-RR LH", "767496296R", "TPV",
     "Santoprene 103-40 (PMR:UB02b,UB05d)", "CELANESE 塞拉尼斯", "TBD",
     "技术协议_20260615.docx; Part List; 报价表-深远", "密度0.9"],
    ["GEA 轮罩", "后右挡泥皮总成 767483010R", "气坝后右 SPOILER-RR RH", "767483010R", "TPV",
     "Santoprene 103-40 (PMR:UB02b,UB05d)", "CELANESE 塞拉尼斯", "TBD",
     "技术协议_20260615.docx; Part List; 报价表-深远", "密度0.9"],
    # 镜面覆盖塑料件 (总成,客户:上海金榕/Mopar-FCA)
    ["镜面覆盖塑料件", "镜面保护罩总成 82219786AA", "镜面罩本体 (LH+RH各1)", "82219786AA", "ASA / ABS",
     "ASA BLACK / ABS (纹理 COD.001)", "未指定(深远拟科思创或国产ASA)", "黑",
     "Mirror protector - FINAL.docx", "客户仅给材料大类,无厂家牌号;整套2件;0.21kg/套"],
    ["镜面覆盖塑料件", "F16 气坝Spat总成 82219733AA", "Spat 气坝本体 (LH+RH)", "82219733AA", "TPO 热塑性聚烯烃",
     "MS.50041 TPO, Type:C (旧号 MS-DC-256), CPN4981", "未指定(FCA材料规格 MS.50041)", "primed底涂/可喷车身色",
     "F16 End Cap_Spat tech sheet_5-18-2026.pptx", "喷漆在墨西哥定制车间,不含"],
    ["镜面覆盖塑料件", "EJ门袋绑带总成 (25/27EJ)", "带扣 Buckle ×4", "—", "ASA (深远判定)",
     "材料规格 MS.50042; 色规 MS-JP-1-3", "未指定", "—",
     "EJ Door bags DVP.xlsx; 27EJ Door Bags-tech sheet.pptx", "MS.50042为FCA材料标准号,深远拟用ASA"],
]

# ---------- Tab2: 总成其余子件 ----------
COLS2 = ["项目", "所属总成", "子件名称", "零件号", "材料/类型",
         "牌号/规格", "厂家/供应商", "类别", "数据来源文件", "备注"]
rows2 = [
    # GEA 挡泥板本体(毛毡压制,非注塑)
    ["GEA 轮罩", "左前挡泥皮总成 638431477R (PHEV/HEV)", "挡泥板本体 MAIN BODY", "—", "PP-PET 聚丙烯+涤纶毛毡",
     "PP-PET 1200gsm", "金榕 JinRong", "压制件(非注塑)", "Part List_20260615.xlsx", "毛毡热压;无指定牌号"],
    ["GEA 轮罩", "右前挡泥皮总成 638422166R (PHEV/HEV)", "挡泥板本体 MAIN BODY", "—", "PP-PET 毛毡",
     "PP-PET 1200gsm", "金榕 JinRong", "压制件(非注塑)", "Part List_20260615.xlsx", "毛毡热压"],
    ["GEA 轮罩", "左前挡泥皮总成 638439763R (BEV)", "挡泥皮本体 MAIN BODY", "—", "PP-PET 毛毡",
     "PP-PET 1200gsm", "金榕 JinRong", "压制件(非注塑)", "Part List_20260615.xlsx", "毛毡热压"],
    ["GEA 轮罩", "右前挡泥皮总成 638424366R (BEV)", "挡泥皮本体 MAIN BODY", "—", "PP-PET 毛毡",
     "PP-PET 1200gsm", "金榕 JinRong", "压制件(非注塑)", "Part List_20260615.xlsx", "毛毡热压"],
    ["GEA 轮罩", "后左挡泥皮总成 767496296R", "挡泥板本体 MAIN BODY", "—", "PP-PET 毛毡",
     "PP-PET 1200gsm", "金榕 JinRong", "压制件(非注塑)", "Part List_20260615.xlsx", "毛毡热压"],
    ["GEA 轮罩", "后右挡泥皮总成 767483010R", "挡泥板本体 MAIN BODY", "—", "PP-PET 毛毡",
     "PP-PET 1200gsm", "金榕 JinRong", "压制件(非注塑)", "Part List_20260615.xlsx", "毛毡热压"],
    # GEA 五金标准件
    ["GEA 轮罩", "各挡泥皮总成(每套×3)", "抽芯铆钉 RIVET φ4*11 + 垫片 φ12*φ5*0.8", "EXT31021028", "金属",
     "RIVET φ4*11 + GASKET φ12*φ5*0.8", "—", "金属标准件", "Part List_20260615.xlsx", "标准件"],
    ["GEA 轮罩", "各挡泥皮总成", "码钉 CODE PIN", "—", "金属", "CODE PIN", "—", "金属标准件",
     "Part List_20260615.xlsx", "6.15新增"],
    # 镜面保护罩 子件
    ["镜面覆盖塑料件", "镜面保护罩总成 82219786AA", "结构胶 Adhesive", "—", "环氧胶 Epoxy",
     "EP-20 / H-97M", "未指定", "采购(涂胶装配)", "Mirror protector - FINAL.docx; DATA SHEET Adhesive.pdf", "非成型件"],
    ["镜面覆盖塑料件", "镜面保护罩总成 82219786AA", "说明书 I-sheet", "—", "纸质/印刷",
     "I-sheet", "—", "采购", "总成报价表(深远)", "QR码由甲方提供"],
    # F16 Spat 子件
    ["镜面覆盖塑料件", "F16 气坝Spat总成 82219733AA", "Bracket Support 支架", "—", "PP基TPO(注塑)",
     "MS.50041 LBI's Hifax TYC 1168P, CPN4998", "LyondellBasell 利安德巴塞尔", "采购(direct buy)",
     "F16 End Cap_Spat tech sheet.pptx", "注塑塑料件,但现有件定向采购,未自制开模"],
    ["镜面覆盖塑料件", "F16 气坝Spat总成 82219733AA", "Sensor Bracket 传感器支架", "—", "PP基TPO(注塑)",
     "MS.50041 LBI's Hifax TYC 1168P, CPN4981", "LyondellBasell 利安德巴塞尔", "采购(direct buy)",
     "F16 End Cap_Spat tech sheet.pptx", "注塑塑料件,定向采购"],
    ["镜面覆盖塑料件", "F16 气坝Spat总成 82219733AA", "U-Clip 卡扣 ×4", "06509800AA", "塑料卡扣",
     "M-4.8 U-Clip (PS-11036)", "未指定", "采购", "F16 End Cap_Spat tech sheet.pptx", "标准紧固卡扣"],
    ["镜面覆盖塑料件", "F16 气坝Spat总成 82219733AA", "垫片 Spacer ×2", "—", "—",
     "Spacer", "未指定", "采购", "总成报价表(深远)", "—"],
    ["镜面覆盖塑料件", "F16 气坝Spat总成 82219733AA", "说明书 I-sheet", "—", "纸质/印刷",
     "I-sheet", "—", "采购", "F16 End Cap_Spat tech sheet.pptx", "QR码由甲方提供"],
    # EJ 门袋 子件
    ["镜面覆盖塑料件", "EJ门袋绑带总成 (25/27EJ)", "尼龙绑带 Nylon Cords/Straps ×2", "—", "尼龙纺织带",
     "材料规格 MS-JZ-12-20 (410mm)", "未指定", "采购(纺织件)", "EJ Door bags DVP.xlsx; 27EJ tech sheet.pptx", "待询价"],
    ["镜面覆盖塑料件", "EJ门袋绑带总成 (25/27EJ)", "紧固件 Fastener ×4", "—", "金属",
     "M8*25", "未指定", "金属标准件", "27EJ Door Bags-tech sheet.pptx (KIT BOM)", "—"],
    ["镜面覆盖塑料件", "EJ门袋绑带总成 (25/27EJ)", "说明书 I-sheet", "—", "纸质/印刷",
     "I-sheet", "—", "采购", "27EJ Door Bags-tech sheet.pptx", "—"],
]

# ---------- 样式 ----------
hdr_fill = PatternFill("solid", fgColor="305496")
hdr_font = Font(bold=True, color="FFFFFF", size=11)
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
proj_fills = {"汽车大灯 FW_P06": "FCE4D6", "GEA 轮罩": "E2EFDA", "镜面覆盖塑料件": "DDEBF7"}


def build_sheet(ws, cols, rows, grade_col):
    ws.append(cols)
    for c in range(1, len(cols) + 1):
        cell = ws.cell(1, c)
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    for r in rows:
        ws.append(r)
        ridx = ws.max_row
        fill = PatternFill("solid", fgColor=proj_fills.get(r[0], "FFFFFF"))
        for c in range(1, len(cols) + 1):
            cell = ws.cell(ridx, c)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = border; cell.fill = fill
            if c == grade_col:
                cell.font = Font(bold=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{ws.max_row}"
    ws.row_dimensions[1].height = 30


wb = openpyxl.Workbook()
ws1 = wb.active
ws1.title = "注塑原材料"
build_sheet(ws1, COLS1, rows1, grade_col=6)
w1 = [16, 32, 22, 13, 22, 34, 26, 16, 42, 34]
for i, w in enumerate(w1, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

ws2 = wb.create_sheet("总成其余子件")
build_sheet(ws2, COLS2, rows2, grade_col=6)
w2 = [16, 32, 30, 16, 22, 34, 26, 18, 42, 28]
for i, w in enumerate(w2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

wb.save(OUT)
print("已保存:", OUT)
print("Tab1 注塑原材料:", len(rows1), "行")
print("Tab2 总成其余子件:", len(rows2), "行")
