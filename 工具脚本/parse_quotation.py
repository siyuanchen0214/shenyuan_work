"""
注塑报价 Excel 解析脚本
用法: python3 parse_quotation.py <excel文件路径>
输出: 打印每个 Sheet 摘要 + 同目录生成 .json
"""

import sys
import json
import openpyxl
from pathlib import Path


def parse_sheet(ws) -> dict:
    def v(row, col):
        val = ws.cell(row=row, column=col).value
        return val if val is not None else None

    def n(row, col):
        try:
            return float(ws.cell(row=row, column=col).value)
        except (TypeError, ValueError):
            return None

    r = {}

    # Header
    r["零件名称"]      = v(4, 15)
    r["零件号"]        = v(5, 15)
    r["成品重量_kg"]   = n(4, 21)
    r["毛坯重量_kg"]   = n(5, 21)
    r["总采购量_万件"] = n(6, 15)
    r["能耗单价_元kwh"]= n(7, 21)
    r["废品率"]        = n(8, 15)
    r["年生产天数"]    = n(8, 21)

    # 一、材料费 (data rows 10-26, total row 27)
    materials = []
    for row in range(10, 27):
        name  = v(row, 5)
        spec  = v(row, 6)
        qty   = n(row, 8)
        rate  = n(row, 9)
        price = n(row, 10)
        cost  = n(row, 11)
        if isinstance(name, str) and name not in ("材料名称",):
            materials.append({
                "材料名称":    name,
                "规格":        str(spec).replace("\n", " ") if spec else None,
                "净用量_kg":   qty,
                "耗用率":      rate,
                "单价_元kg":   price,
                "材料费_元件": cost,
            })
    r["材料"] = materials
    r["材料费合计"] = n(27, 11)

    # 三、模具摊销 (data rows 30-39, total row 40)
    molds = []
    for row in range(30, 40):
        name  = v(row, 5)
        price = n(row, 6)
        cav   = n(row, 7)
        life  = n(row, 8)
        alloc = n(row, 9)
        unit  = n(row, 11)
        if isinstance(name, str):
            molds.append({
                "模具名称":      name,
                "模具价格_万元": price,
                "穴数":          cav,
                "寿命_万模":     life,
                "分摊数量_万件": alloc,
                "单件分摊_元":   unit,
            })
    r["模具"] = molds
    r["模具摊销合计"] = n(40, 11)

    # 五、加工工序 (data rows 42-57, total row 58)
    processes = []
    for row in range(42, 58):
        equip   = v(row, 5)
        cycle   = n(row, 6)
        tau     = n(row, 8)
        workers = n(row, 9)
        wage    = n(row, 10)
        oee     = n(row, 12)
        labor   = n(row, 14)
        eq_val  = n(row, 16)
        depre   = n(row, 18)
        power   = n(row, 19)
        util    = n(row, 20)
        energy  = n(row, 21)
        if isinstance(equip, str):
            processes.append({
                "设备":          equip,
                "周期_秒":       cycle,
                "单件台时_h":    tau,
                "操作人数":      workers,
                "工时单价_元h":  wage,
                "稼动率":        oee,
                "人工费_元件":   labor,
                "设备价_万元":   eq_val,
                "折旧_元件":     depre,
                "功率_kw":       power,
                "能耗利用率":    util,
                "能耗费_元件":   energy,
            })
    r["工序"] = processes
    r["人工合计"] = n(58, 14)
    r["折旧合计"] = n(58, 18)
    r["能耗合计"] = n(58, 21)

    # 汇总
    r["直接工资"] = n(59, 5)
    r["废品费用"] = n(60, 5)
    r["管理费用"] = n(61, 5)
    r["运输费用"] = n(62, 5)
    r["包装费用"] = n(63, 5)
    r["成本价"]   = n(64, 5)
    r["利润率"]   = n(65, 5)
    r["利润"]     = n(66, 5)
    r["未税报价"] = n(67, 5)
    r["增值税13%"]= n(68, 5)
    r["含税总价"] = n(69, 5)

    return r


def print_summary(name: str, d: dict):
    print(f"\n{'='*52}")
    print(f"  {name}  |  零件号: {d.get('零件号')}")
    print(f"  成品重量: {d.get('成品重量_kg')} kg  |  废品率: {d.get('废品率')}")
    print(f"{'='*52}")
    for m in d.get("材料", []):
        print(f"  [材料] {m['材料名称']}  净用量={m['净用量_kg']}kg  "
              f"单价={m['单价_元kg']}元/kg  → {m['材料费_元件']}元/件")
    print(f"  材料费合计:   {d.get('材料费合计')} 元/件")
    for md in d.get("模具", []):
        print(f"  [模具] {md['模具名称']}  {md['模具价格_万元']}万  "
              f"{md['穴数']}穴  寿命{md['寿命_万模']}万模  → {md['单件分摊_元']}元/件")
    print(f"  模具摊销合计: {d.get('模具摊销合计')} 元/件")
    for p in d.get("工序", []):
        print(f"  [工序] {p['设备']}  周期{p['周期_秒']}s  "
              f"人工{p['人工费_元件']} + 折旧{p['折旧_元件']} + 能耗{p['能耗费_元件']} 元/件")
    print(f"  ── 成本价: {d.get('成本价')} 元/件")
    print(f"  ── 利润率: {d.get('利润率')}")
    print(f"  ── 含税总价: {d.get('含税总价')} 元/件")


def parse_file(path: str) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True)
    skip = {"Sheet1", "Sheet2"}
    results = {}
    for name in wb.sheetnames:
        if name.strip() in skip:
            continue
        data = parse_sheet(wb[name])
        results[name] = data
        print_summary(name, data)
    return results


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("用法: python3 parse_quotation.py <excel文件路径>")
        sys.exit(1)
    fpath = sys.argv[1]
    results = parse_file(fpath)
    out = Path(fpath).with_suffix(".json")
    out.write_text(json.dumps(results, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\n完整数据已输出: {out}")
