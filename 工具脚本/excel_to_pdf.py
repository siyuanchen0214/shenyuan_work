#!/usr/bin/env python3
"""
excel_to_pdf.py — 将 Excel 文件每个可见 tab 压缩为横向单页 PDF

用法:
    python3 excel_to_pdf.py <文件1.xlsx> [文件2.xlsx ...]
    python3 excel_to_pdf.py /path/to/folder/          # 转换文件夹内所有 xlsx

规则:
    1. 每个可见 tab 横向 fit 到 1 页
    2. 原始 xlsx 文件不会被修改
    3. PDF 输出在同一目录，文件名和 xlsx 相同（扩展名换 .pdf）
    4. 转换完成后自动验证：原文件时间戳未变 + PDF 页数 == 可见 tab 数
"""

import sys
import os
import glob
import shutil
import tempfile
import time

import openpyxl
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from openpyxl.utils import get_column_letter

try:
    import xlwings as xw
except ImportError:
    sys.exit("缺少 xlwings，请先运行: pip3 install xlwings")

try:
    from pypdf import PdfReader
except ImportError:
    sys.exit("缺少 pypdf，请先运行: pip3 install pypdf")


def collect_files(args):
    paths = []
    for arg in args:
        if os.path.isdir(arg):
            paths += glob.glob(os.path.join(arg, "*.xlsx"))
        elif os.path.isfile(arg) and arg.endswith(".xlsx"):
            paths.append(arg)
        else:
            print(f"  跳过（不是 xlsx 或不存在）: {arg}")
    # 过滤掉临时锁文件（~$ 开头）
    return [p for p in paths if not os.path.basename(p).startswith("~$")]


def get_content_range(ws):
    """返回 sheet 中所有有内容的最小/最大行列（包含合并单元格）"""
    min_row = min_col = float('inf')
    max_row = max_col = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                min_row = min(min_row, cell.row)
                min_col = min(min_col, cell.column)
                max_row = max(max_row, cell.row)
                max_col = max(max_col, cell.column)
    for merge in ws.merged_cells.ranges:
        min_row = min(min_row, merge.min_row)
        min_col = min(min_col, merge.min_col)
        max_row = max(max_row, merge.max_row)
        max_col = max(max_col, merge.max_col)
    if max_row == 0:
        return None
    return min_row, min_col, max_row, max_col


def prepare_temp(xlsx_path, tmp_path):
    """复制 xlsx 到临时文件，扩展 print area 并设置 fit-to-page（横向）"""
    wb = openpyxl.load_workbook(xlsx_path)
    visible_count = 0

    for ws in wb.worksheets:
        if ws.sheet_state == 'hidden':
            continue
        visible_count += 1

        rng = get_content_range(ws)
        if rng is None:
            continue
        min_row, min_col, max_row, max_col = rng

        # 把 print area 扩展到全部有内容的区域
        min_col_letter = get_column_letter(min_col)
        max_col_letter = get_column_letter(max_col)
        ws.print_area = f"${min_col_letter}${min_row}:${max_col_letter}${max_row}"

        # 横向 + fit to 1×1 page
        ws.page_setup.orientation = 'landscape'
        ws.page_setup.fitToWidth  = 1
        ws.page_setup.fitToHeight = 1

        if ws.sheet_properties is None:
            ws.sheet_properties = WorksheetProperties()
        if ws.sheet_properties.pageSetUpPr is None:
            ws.sheet_properties.pageSetUpPr = PageSetupProperties()
        ws.sheet_properties.pageSetUpPr.fitToPage = True

    wb.save(tmp_path)
    return visible_count


def convert_one(xlsx_path, app):
    pdf_path = os.path.splitext(xlsx_path)[0] + ".pdf"
    tmp_fd, tmp_path = tempfile.mkstemp(suffix=".xlsx", prefix="_tmp_")
    os.close(tmp_fd)

    try:
        # 1. 记录原文件修改时间
        orig_mtime = os.path.getmtime(xlsx_path)

        # 2. 准备临时文件（带 fit-to-page 设置）
        visible_count = prepare_temp(xlsx_path, tmp_path)

        # 3. 用 xlwings 导出 PDF（不碰原文件）
        wb = app.books.open(tmp_path)
        wb.to_pdf(path=pdf_path)
        wb.close()

    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)

    # ── 验证 1：原文件未被修改 ──────────────────────────
    new_mtime = os.path.getmtime(xlsx_path)
    orig_safe = abs(new_mtime - orig_mtime) < 1  # 容差 1 秒

    # ── 验证 2：PDF 页数 == 可见 tab 数 ─────────────────
    try:
        pdf_pages = len(PdfReader(pdf_path).pages)
    except Exception:
        pdf_pages = -1
    pages_ok = (pdf_pages == visible_count)

    return {
        "xlsx":          os.path.basename(xlsx_path),
        "pdf":           pdf_path,
        "pdf_pages":     pdf_pages,
        "visible_tabs":  visible_count,
        "orig_safe":     orig_safe,
        "pages_ok":      pages_ok,
        "size_kb":       os.path.getsize(pdf_path) // 1024 if os.path.exists(pdf_path) else 0,
    }


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(0)

    files = collect_files(sys.argv[1:])
    if not files:
        sys.exit("没有找到可处理的 xlsx 文件")

    print(f"\n共找到 {len(files)} 个文件，开始转换...\n")

    app = xw.App(visible=False)
    results = []
    try:
        for f in files:
            print(f"  处理: {os.path.basename(f)}")
            try:
                r = convert_one(f, app)
                results.append(r)
            except Exception as e:
                print(f"    ✗ 出错: {e}")
                results.append({"xlsx": os.path.basename(f), "error": str(e)})
    finally:
        app.quit()

    # ── 汇总报告 ─────────────────────────────────────────
    print("\n" + "="*62)
    print(f"{'文件':<38} {'原文件':<6} {'页数':<10} {'大小'}")
    print("-"*62)
    all_ok = True
    for r in results:
        if "error" in r:
            print(f"  {r['xlsx'][:36]:<38} ✗ 出错: {r['error']}")
            all_ok = False
            continue
        orig_mark  = "✓ 未改" if r["orig_safe"]  else "✗ 被改!"
        pages_mark = f"✓ {r['pdf_pages']}页/{r['visible_tabs']}tab" if r["pages_ok"] else f"✗ {r['pdf_pages']}页≠{r['visible_tabs']}tab"
        print(f"  {r['xlsx'][:36]:<38} {orig_mark:<6} {pages_mark:<12} {r['size_kb']}KB")
        if not r["orig_safe"] or not r["pages_ok"]:
            all_ok = False

    print("="*62)
    print("全部通过 ✓" if all_ok else "有问题，请检查上方标 ✗ 的行")
    print()


if __name__ == "__main__":
    main()
